# Hands-on with Python and OpenTelemetry
# This is for educational purpose; Scrapy et al is not used

# https://stackoverflow.com/questions/231767/what-does-the-yield-keyword-do
# https://tinydb.readthedocs.io/en/latest/index.html
# https://regex101.com

import argparse
import requests
import re
import bs4
import json
import glob
from openpyxl import Workbook
import datetime

class REObject:
    def __init__(self):
        self.url = ''
        self.title = ''
        self.willhaben_id = ''
        self.price = -1
        self.squarefeet = -1
        #self.rooms = -1
        #self.agency = ''
        #first_seen = datetime
        
    def from_json(self, json):
        self.url = json['url']
        self.title = json['title']
        self.willhaben_id = json['willhaben_id']
        self.price = json['price']
        self.squarefeet = json['squarefeet']
    
    def __str__(self):
        return "REObject: % s\n% s EUR, % sm2, willhaben_id: %s" % (self.title, self.price, self.squarefeet, self.willhaben_id)
        
    # convert string containing EUR currency to int (hack, to be improved)
    def string_to_int(str):
        result = 0
        
        if str is not None:
            if ',' in str:
                str = str[0:str.rfind(',')] # ignore beyond first ','
            if '.' in str:
                str = str.replace('.', '') # remove all '.'s
            result = int(str)
        
        return result

class Urtyp83:
    # constants
    BASE_URL = 'https://www.willhaben.at/'
    START_URL = 'https://www.willhaben.at/iad/immobilien/mietwohnungen/oberoesterreich/linz?page=1&rows=100'
    #START_URL = 'https://www.willhaben.at/iad/immobilien/mietwohnungen/oberoesterreich/linz?page=46'
    MAX_RESULTS = 2000

    # instance members
    next_page = START_URL
    
    # crawl a single real estate page
    def crawl_re(url):
        res = requests.get(url)
        if res.status_code == 200:
            html = bs4.BeautifulSoup(res.text, 'html.parser')
            
            reobject = REObject()
            reobject.url = url
            reobject.title = html.title.contents[0]
            
            willhaben_id = re.findall('-\d+\/$', url)[0]
            reobject.willhaben_id = willhaben_id[1:-1]
            
            """ soup = html.find_all('div', {'data-testid': 'contact-box-price-box'})
            for s in soup:
                str = re.findall('\d+[\.|,]?\d*[\.|,]?\d*', s.text)[0]
                i = string_to_int(str)
                print(f'price: {i}') """
                    
            soup = html.find_all(text=re.compile('Gesamtmiete inkl. MWSt'))
            if len(soup) > 0:
                str = re.findall('\d+[\.|,]?\d*[\.|,]?\d*', soup[0].parent.parent.text)[0]
                reobject.price = REObject.string_to_int(str)

            soup = html.find_all(text=re.compile('Wohnfläche'))
            if len(soup) > 0:
                    str = re.findall('\d*[,\d*]', soup[0].parent.parent.next_sibling.text)
                    if str is not None:
                        reobject.squarefeet = REObject.string_to_int(str[0])
            else:
                soup = html.find_all(text=re.compile('Nutzfläche'))
                if len(soup) > 0:
                    str = re.findall('\d*[,\d*]', soup[0].parent.parent.next_sibling.text)[0]
                    reobject.squarefeet = REObject.string_to_int(str)
                
            return reobject

    # get next page (pagination)
    def get_next_page(res):
        html = bs4.BeautifulSoup(res.text, 'html.parser')

        # get total number or search results
        total_results = html.find_all('h1', {'data-testid': 'result-list-title'})
        total = re.findall('\d*', total_results[0].text.replace('.', ''))[0]
        print(f'Total number of listings: {total}')
        # https://github.com/maksimKorzh/scrapy-tutorials/blob/master/src/willhaben/willhaben.py
        
        # return the URL of the 'next' page
        try:
            soup = html.find_all('a', {'data-testid': 'pagination-top-next-button'})
            for s in soup:
                print(s['href'])
                return s['href']
        except:
            return None

    # stats
    def print_stats(reobjects, url_errors=None):
        total_price = 0
        total_squarefeet = 0
        
        for reobject in reobjects:
            total_price = total_price + reobject.price
            total_squarefeet = total_squarefeet + reobject.squarefeet
        
        no_objects = len(reobjects)
        
        print(f'Pages crawled: {no_objects}')

        print(f'Avg price: {total_price/no_objects}')
        print(f'Avg m2: {total_squarefeet/no_objects}')
        print(f'Price per m2: EUR {total_price/total_squarefeet}')

        if url_errors != None and url_errors > 0:
            print(f'Errors:{len(url_errors)}')
            for url_error in url_errors:
                print(url_error)

    # write xls
    def write_speadsheet(reobjects, output_dir):
        wb = Workbook()
        ws = wb.create_sheet('willhaben_crawler', 0) # new sheet is first tab
        
        ws['A1'] = "Preis inkl. MWst"
        ws['B1'] = "m2"
        ws['C1'] = "Preis pro m2"
        ws['D1'] = "Beschreibung"
        ws['E1'] = "Link auf willhaben.at"

        row = 2
        
        for reobject in reobjects:
            ws["A"+str(row)] = reobject.price
            ws["B"+str(row)] = reobject.squarefeet
            if(reobject.squarefeet > 0):
                ws["C"+str(row)] = reobject.price / reobject.squarefeet
            ws["D"+str(row)] = reobject.title
            ws["E"+str(row)] = reobject.url
            row = row + 1
        
        d = datetime.datetime.now()
        fname = f'{output_dir}/willhaben_{d.year}-{d.month}-{d.day}_{d.hour}-{d.minute}.xlsx'
        wb.save(fname)
        return fname

    # write json
    def write_json(reobjects, output_dir):
        d = datetime.datetime.now()
        fname = f'{output_dir}/willhaben_{d.year}-{d.month}-{d.day}_{d.hour}-{d.minute}.json'
        json_file = open(fname, 'w')
        
        j = json.dumps(reobjects, default=vars)
        
        json_file.write(j)
        return fname
    
    # read json
    def read_json(filename):
        f = open(filename)
        data = json.load(f)
        
        reobjects = []

        for d in data:
            reobject = REObject()
            reobject.from_json(d)
            reobjects.append(reobject)
            
        return reobjects

    # write html, for debug purposes
    def write_html(html, fname):
        html_file = open(fname, 'wb')

        for chunk in html.iter_content(100000):
            html_file.write(chunk)

        html_file.close()
        print(f'File written: {html_file.name}')

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=f'Scrape real estate listings from {Urtyp83.START_URL}')
    parser.add_argument('-v', '--verbose', action='store_true', help='verbose mode')
    parser.add_argument('--xls', action='store_true', help='store results in willhaben.xlsx')
    parser.add_argument('--json', action='store_true', help='store results in json willhaben.json')
    parser.add_argument('--dir', help='relative path to input/output directory')
    parser.add_argument('--read', action='store_true', help='read json files from --dir')
    args = parser.parse_args()
    
    inout_dir = "."
    
    if args.dir != None:
        inout_dir = args.dir
        
    if args.read:

        json_filenames = []
        json_timestamps = []
                
        for f in glob.glob('*.json'): # TODO inout_dir
            json_filenames.append(f)
        
        formatter = '%Y-%m-%d_%H-%M'
        for f in json_filenames:
            f = f.replace('willhaben_', '')
            f = f.replace('.json', '')
            timestamp = datetime.datetime.strptime(f, formatter)
            json_timestamps.append(timestamp)
        
        most_recent = json_timestamps[0]
        for i in range(len(json_filenames)):
            #if json_timestamps[i]
            print(json_filenames[i])
            print(json_timestamps[i])
            print('---')
        
        exit(1)
            
        reobj = Urtyp83.read_json(file)
        Urtyp83.print_stats(reobj)
        print("---")

        
        file_1 = 'willhaben_2021-11-29_23-44.json'
        file_2 = 'willhaben_2021-11-30_18-35.json'
        
        file_1 = file_1.replace('willhaben_', '')
        file_1 = file_1.replace('.json', '')
        print(file_1)
        

        print(timestamp)
        exit(1)
        
        reobj_1 = Urtyp83.read_json(file_1)
        reobj_2 = Urtyp83.read_json(file_2)
        
        Urtyp83.print_stats(reobj_1)
        Urtyp83.print_stats(reobj_2)
        print('---')
        
        found = []
        for reobj in reobj_2:
            for tmp in reobj_1:
                if tmp.willhaben_id == reobj.willhaben_id:
                    found.append(tmp)
                    break
        
        print(f'From {len(reobj_1)} properties, {len(reobj_2)-len(found)} have been added.')
        exit(1)
    
    ur = Urtyp83()
    
    reobjects = []
    url_errors = []

    i = 0
    next_url = Urtyp83.START_URL
    res = None
    
    ts_start = datetime.datetime.now()
    
    while i < Urtyp83.MAX_RESULTS:

        if res is None:
            res = requests.get(next_url)
            if res.status_code != 200:
                print(f'Could not parse {next_url}')
                exit(1)

            urls = re.findall('\"\/iad\/immobilien\/d\/.*?\/\"', res.text) 
            
            for u in urls:
                url = Urtyp83.BASE_URL + u[1:-1]
                try:
                    reobject = Urtyp83.crawl_re(url)

                    print(f'{i}: {url}')
                    reobjects.append(reobject)
                    print(reobject)
                    print('---')
                    
                except:
                    url_errors.append(url)

                i += 1
                if i >= Urtyp83.MAX_RESULTS:
                    break

            tmp = Urtyp83.get_next_page(res)
            if tmp is None:
                break
            
            next_url = Urtyp83.BASE_URL + tmp
            res = None
        
    print(f'DONE: {len(reobjects)} properties crawled [took {datetime.datetime.now()-ts_start}]')
    
    if args.xls:
        print(f'File written: {Urtyp83.write_speadsheet(reobjects, inout_dir)}')

    if args.json:
        print(f'File written: {Urtyp83.write_json(reobjects, inout_dir)}')